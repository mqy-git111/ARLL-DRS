# Copyright 2020 - 2022 MONAI Consortium
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#     http://www.apache.org/licenses/LICENSE-2.0
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import argparse
import json
import os
import openpyxl

import nibabel as nib
import numpy as np
import torch
from utils.data_utils import get_loader
from utils.utils import dice, resample_3d

from monai.inferers import sliding_window_inference
from monai.networks.nets import SwinUNETR

parser = argparse.ArgumentParser(description="Swin UNETR segmentation pipeline")
parser.add_argument(
    "--pretrained_dir", default="./result/pelvic_swin_2/", type=str, help="pretrained checkpoint directory"
)
parser.add_argument("--data_dir", default="./dataset/", type=str, help="dataset directory")
parser.add_argument("--exp_name", default="pelvic_swin_2/", type=str, help="experiment name")
parser.add_argument("--json_list", default="dataset_pelvic.json", type=str, help="dataset json file")
parser.add_argument(
    "--pretrained_model_name",
    default="model_234.pt",
    type=str,
    help="pretrained model name",
)
parser.add_argument("--feature_size", default=48, type=int, help="feature size")
parser.add_argument("--infer_overlap", default=0.5, type=float, help="sliding window inference overlap")
parser.add_argument("--in_channels", default=1, type=int, help="number of input channels")
parser.add_argument("--out_channels", default=25, type=int, help="number of output channels")
parser.add_argument("--a_min", default=-1019.0, type=float, help="a_min in ScaleIntensityRanged")
parser.add_argument("--a_max", default=2230.0, type=float, help="a_max in ScaleIntensityRanged")
parser.add_argument("--b_min", default=0.0, type=float, help="b_min in ScaleIntensityRanged")
parser.add_argument("--b_max", default=1.0, type=float, help="b_max in ScaleIntensityRanged")
parser.add_argument("--space_x", default=1.5, type=float, help="spacing in x direction")
parser.add_argument("--space_y", default=1.5, type=float, help="spacing in y direction")
parser.add_argument("--space_z", default=3.0, type=float, help="spacing in z direction")
parser.add_argument("--roi_x", default=96, type=int, help="roi size in x direction")
parser.add_argument("--roi_y", default=96, type=int, help="roi size in y direction")
parser.add_argument("--roi_z", default=96, type=int, help="roi size in z direction")
parser.add_argument("--dropout_rate", default=0.0, type=float, help="dropout rate")
parser.add_argument("--distributed", action="store_true", help="start distributed training")
parser.add_argument("--workers", default=1, type=int, help="number of workers")
parser.add_argument("--RandFlipd_prob", default=0.2, type=float, help="RandFlipd aug probability")
parser.add_argument("--RandRotate90d_prob", default=0.2, type=float, help="RandRotate90d aug probability")
parser.add_argument("--RandScaleIntensityd_prob", default=0.1, type=float, help="RandScaleIntensityd aug probability")
parser.add_argument("--RandShiftIntensityd_prob", default=0.1, type=float, help="RandShiftIntensityd aug probability")
parser.add_argument("--spatial_dims", default=3, type=int, help="spatial dimension of input data")
parser.add_argument("--use_checkpoint", action="store_true", help="use gradient checkpointing to save memory")
def read_json_file(file_path):
    with open(file_path, 'r') as file:
        data = json.load(file)
        return data

def main():
    args = parser.parse_args()
    args.test_mode = True
    #将数据写入xlsx
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    en_organ = []
    ch_organ = []
    json_file = read_json_file(os.path.join(args.data_dir,args.json_list))
    organ_dict_en = json_file["labels"]
    organ_dict_ch = json_file["labels_ch"]
    for k in organ_dict_en:
        en_organ.append(organ_dict_en[k])
    for k in organ_dict_ch:
        ch_organ.append(organ_dict_ch[k])
    col_num = 1
    for row, value in enumerate(en_organ, start=1):
        sheet.cell(row=row, column=col_num, value=value)
    col_num = 2
    for row, value in enumerate(ch_organ, start=1):
        sheet.cell(row=row, column=col_num, value=value)
    output_directory = "./outputs/" + args.exp_name
    predict_path = os.path.join(output_directory,"predict")
    # img_path = os.path.join(output_directory,"image")
    # label_path = os.path.join(output_directory,"label")
    if not os.path.exists(predict_path):
        os.makedirs(predict_path)
    # if not os.path.exists(img_path):
    #     os.makedirs(img_path)
    # if not os.path.exists(label_path):
    #     os.makedirs(label_path)
    val_loader = get_loader(args)
    pretrained_dir = args.pretrained_dir
    model_name = args.pretrained_model_name
    # device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    pretrained_pth = os.path.join(pretrained_dir, model_name)
    model = SwinUNETR(
        img_size=96,
        in_channels=args.in_channels,
        out_channels=args.out_channels,
        feature_size=args.feature_size,
        drop_rate=0.0,
        attn_drop_rate=0.0,
        dropout_path_rate=0.0,
        use_checkpoint=args.use_checkpoint,
    )
    model_dict = torch.load(pretrained_pth)["state_dict"]
    model.load_state_dict(model_dict)
    model.eval()
    model.to(device)

    with torch.no_grad():
        count = [0 for i in range(0, args.out_channels - 1)]
        dice_all = [0 for i in range(0, args.out_channels - 1)]
        dice_avg = [0 for i in range(0, args.out_channels - 1)]
        # dice_list_case = []
        col_num = 3
        for i, batch in enumerate(val_loader):
            val_inputs, val_labels = (batch["image"].cuda(), batch["label"].cuda())
            original_affine = batch["label_meta_dict"]["affine"][0].numpy()
            _, _, h, w, d = val_labels.shape
            target_shape = (h, w, d)
            img_name = batch["image_meta_dict"]["filename_or_obj"][0].split("/")[-1]
            data = [img_name[:-7]]
            # print("Inference on case {}".format(img_name))
            val_outputs = sliding_window_inference(
                val_inputs, (args.roi_x, args.roi_y, args.roi_z), 4, model, overlap=args.infer_overlap, mode="gaussian"
            )
            val_outputs = torch.softmax(val_outputs, 1).cpu().numpy()
            val_outputs = np.argmax(val_outputs, axis=1).astype(np.uint8)[0]
            val_labels = val_labels.cpu().numpy()[0, 0, :, :, :]
            val_outputs = resample_3d(val_outputs, target_shape)
            dice_list_sub = []
            count_temp = 0
            dice_temp = []
            #只计算有标签的数据的dice
            for i in range(1, args.out_channels):
                organ_Dice = dice(val_outputs == i, val_labels == i)
                if (val_labels == i).sum() == 0:
                    dice_all[i - 1] += 0.0
                    data.append("NAN")
                else:
                    # dice_all[i-1]+=dice[i]
                    dice_all[i - 1] += organ_Dice
                    count[i - 1] = count[i - 1] + 1
                    count_temp+=1
                    dice_temp.append(organ_Dice)
                    data.append(organ_Dice)

                dice_list_sub.append(organ_Dice)
            # mean_dice_old = np.mean(dice_list_sub)
            mean_dice_new = np.mean(dice_temp)
            data.append(mean_dice_new)
            # print(img_name + "Mean Organ Dice old: {}".format(mean_dice_old))
            print(img_name + "Mean Organ Dice: {}".format(mean_dice_new))
            # dice_list_case.append(mean_dice_old)
            nib.save(
                nib.Nifti1Image(val_outputs.astype(np.uint8), original_affine), os.path.join(predict_path, img_name)
            )
            # nib.save(
            #     nib.Nifti1Image(val_labels.astype(np.uint8), original_affine), os.path.join(label_path, img_name)
            # )
            # nib.save(
            #     nib.Nifti1Image(val_inputs[0,0,:,:,:].astype(np.float32), original_affine), os.path.join(img_path, img_name)
            # )

            #写入exlx
            for row, value in enumerate(data, start=1):
                sheet.cell(row=row, column=col_num, value=value)
            col_num += 1
            # break
            # if col_num >=10:
            #     break
        #计算平均dice
        for i in range(0, args.out_channels - 1):
            if count[i] == 0:
                dice_avg[i] = 0
            else:
                dice_avg[i] = dice_all[i] / count[i]
        dice_avg.append(np.mean(dice_avg))
        for row, value in enumerate(dice_avg, start=2):
            sheet.cell(row=row, column=col_num, value=value)
        # sheet["A1"] = np.mean(dice_avg)
        workbook.save(os.path.join(output_directory,"predict_234.xlsx"))
        # print("Overall Mean Dice old: {}".format(np.mean(dice_list_case)))
        print(dice_avg)
        print("Overall Mean Dice new: {}".format(np.mean(dice_avg)))


if __name__ == "__main__":
    main()
